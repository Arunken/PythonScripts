# -*- coding: utf-8 -*-
"""
Created on Tue Mar 27 15:17:10 2018

@author: Kenrig
"""


# 2) =======Read excel,Open excel, get sheets from workbook, getting sheets from the sheets
# getting rows and columns from the sheets

from pyexcel_xls import save_data, read_data

data = {"sheet 1":[[1,2,3],[4,5,6]], "sheet 2":[[2,3,'pavan'],['python','data']]}
save_data("myxls.xls",data)

rdata = read_data("myxls.xls")
print(rdata)


import xlsxwriter

workbook = xlsxwriter.Workbook('test.xlsx')
worksheet = workbook.add_worksheet() 
worksheet.write('A2','Test data')
workbook.close()
#============= Answer using xlrd ===================================================================

from __future__ import print_function
from os.path import join, dirname, abspath,__file__
import xlrd

# fname = join(dirname(dirname(abspath(__file__))), 'DataForFiles', 'Superstore.xls')
fname = 'E:\\Documents\\PythonProjects\\1_Basics\\DataForFiles\\Superstore.xls'

# Open the workbook
xl_workbook = xlrd.open_workbook(fname)

# List sheet names, and pull a sheet by name
#
sheet_names = xl_workbook.sheet_names()
print('Sheet Names', sheet_names)

xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])

# Or grab the first sheet by index 
#  (sheets are zero-indexed)
#
xl_sheet = xl_workbook.sheet_by_index(0)
print ('Sheet name: %s' % xl_sheet.name)

# Pull the first row by index
#  (rows/columns are also zero-indexed)
#
row = xl_sheet.row(0)  # 1st row

# Print 1st row values and types
#
from xlrd.sheet import ctype_text   

print('(Column #) type:value')
for idx, cell_obj in enumerate(row):
    cell_type_str = ctype_text.get(cell_obj.ctype, 'unknown type')
    print('(%s) %s %s' % (idx, cell_type_str, cell_obj.value))

# Print all values, iterating through rows and columns
#
num_cols = xl_sheet.ncols   # Number of columns
for row_idx in range(0, xl_sheet.nrows):    # Iterate through rows
    print ('-'*40)
    print ('Row: %s' % row_idx)   # Print row number
    for col_idx in range(0, num_cols):  # Iterate through columns
        cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
        print ('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
        
        
# 1) Perform following on grade analysis data :
import pandas as pd

df= pd.read_excel('E:\\Documents\\PythonProjects\\1_Basics\\DataForFiles\\grades.xlsx') #Reading the dataset in a dataframe using Pandas

pernum = []


for name,values in df.iteritems():

    print('=========')
    print(values)
    print('=========')
    
        

    
per = []

for i in range(2,len(df)+2):
    per.append('=SUM(A'+str(i)+':F'+str(i)+')/'+str(len(df)))
   
    

df['Percentage'] = per

df.to_excel('E:\\Documents\\PythonProjects\\1_Basics\\DataForFiles\\grades.xlsx',index = False) 

writer = pd.ExcelWriter('output.xlsx', engine='xlsxwriter')

df.to_excel(writer, index=False, sheet_name='Sheet1')

workbook = writer.book

worksheet = writer.sheets['Sheet1']
header_fmt = workbook.add_format({'bold': True})
worksheet.set_row(0, None, header_fmt)
writer.save()



import matplotlib.pyplot as plt
# Define a function for a histogram
def histogram(data, x_label, y_label, title):
    _, ax = plt.subplots()
    ax.hist(data, color = '#539caf')
    ax.set_ylabel(y_label)
    ax.set_xlabel(x_label)
    ax.set_title(title)

# Call the function to create plot
histogram(data = per
           , x_label = 'Percentage'
           , y_label = 'Frequency'
           , title = 'Distribution of Registered Check Outs')











# 1) Perform following on grade analysis data :

import openpyxl
from os.path import join, dirname, abspath
#from openpyxl.cell import get_column_letter, column_index_from_string

fname = join(dirname(dirname(abspath(__file__))), 'DataForFiles', 'grades.xlsx')
wb = openpyxl.load_workbook(filename=fname)
ws = wb['Sheet1']


rows =0
for index, row in enumerate(ws.iter_rows()):
    rows=rows+1
    cols=0
    if index>0:
        
        for cell in row:
            print(cell.get_column_letter)
            cols=cols+1
            print(ws.cell(row=index + 1, column=1).value, cell.value)



            

            









