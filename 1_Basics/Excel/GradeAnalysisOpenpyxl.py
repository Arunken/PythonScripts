# -*- coding: utf-8 -*-
"""
Created on Thu May 31 15:36:54 2018

@author: SilverDoe
"""

# 1) Perform following on grade analysis data :


import pandas as pd
import openpyxl
from openpyxl import Workbook , load_workbook

source_file = "E:\\Documents\\PythonProjects\\1_Basics\\DataForFiles\\grades1.xlsx"
# write to file
# =============================================================================
# wb = load_workbook (source_file)
# ws = wb.active
# ws.title = "hello world"
# ws.append ([10,10])
# wb.save(source_file)
# =============================================================================

# read from file
df = pd.read_excel(source_file)
sum_jan = df [0].sum() 
print (sum_jan)



# =============================================================================
import openpyxl
#from os.path import join, dirname, abspath

#fname = join(dirname(dirname(abspath(__file__))), 'DataForFiles', 'grades.xlsx')
fname = 'E:\\Documents\\PythonProjects\\1_Basics\\DataForFiles\\grades1.xlsx'
wb = openpyxl.load_workbook(filename=fname)
ws = wb['Sheet1'] # ws = wb.worksheets[0]


rows =0
for index, row in enumerate(ws.iter_rows()):
    rows=rows+1
    cols=0
    if index>0:
        
        for cell in row:
            
            cols=cols+1
            #print(ws.cell(row=index + 1, column=1).value, cell.value)
            #print(cell.value)
        #print('columns : ',cols)
        #ws.write(index,cols+1,'=SUM()')
        print(cell.coordinate,cell.col_idx)
        if(rows>2):
            ws.cell(row=index, column=cell.col_idx).value = 'hello'
        
wb.save(fname)
            